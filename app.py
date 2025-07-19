from flask import Flask, request, jsonify, send_from_directory
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import WebDriverException
import time
import openpyxl
import os
import requests as http_requests
import json
import re
import traceback
from zipfile import BadZipFile
from urllib.parse import urljoin

GEMINI_API_KEY = 'AIzaSyC0KCgBiKHHKbOvOPMB1mL1PbJsVXmmtF8'
GEMINI_URL = f'https://generativelanguage.googleapis.com/v1/models/gemini-1.5-flash-8b:generateContent?key={GEMINI_API_KEY}'
EXCEL_PATH = os.path.join(os.path.dirname(__file__), 'xl_files', 'job.xlsx')

app = Flask(__name__)
app.config['PROPAGATE_EXCEPTIONS'] = True

def extract_jobs_from_html(driver):
    jobs = []
    selectors = [
        '.job-listing', '.job-item', '.job-card', '.opening', '.position', '.career', '.vacancy',
        'li', 'tr', 'div'
    ]
    for selector in selectors:
        cards = driver.find_elements(By.CSS_SELECTOR, selector)
        for card in cards:
            try:
                title = ''
                location = ''
                desc = ''
                for tsel in ['.title', '.job-title', '.position', '.posting-title', 'h2', 'h3', 'a', 'span']:
                    try:
                        title = card.find_element(By.CSS_SELECTOR, tsel).text
                        if title: break
                    except: continue
                for lsel in ['.location', '.job-location', '.city', 'span', 'div']:
                    try:
                        location = card.find_element(By.CSS_SELECTOR, lsel).text
                        if location: break
                    except: continue
                for dsel in ['.description', '.job-desc', '.desc', '.summary', 'p', 'div']:
                    try:
                        desc = card.find_element(By.CSS_SELECTOR, dsel).text
                        if desc: break
                    except: continue
                if title:
                    jobs.append({'Job': title, 'Location': location, 'Description': desc})
            except: continue
    # Remove duplicates
    seen = set()
    unique_jobs = []
    for job in jobs:
        key = (job['Job'], job['Location'], job['Description'])
        if key not in seen:
            seen.add(key)
            unique_jobs.append(job)
    return unique_jobs

@app.route('/')
def serve_index():
    return send_from_directory('static', 'index.html')

@app.route('/scrape', methods=['POST'])
def scrape():
    step = 'init'
    try:
        data = request.get_json()
        url = data.get('url')
        if not url:
            step = 'get url from request'
            raise ValueError('No URL provided.')
        # Set up Selenium with headless Chrome
        step = 'start selenium'
        chrome_options = Options()
        chrome_options.add_argument('--headless')
        chrome_options.add_argument('--no-sandbox')
        chrome_options.add_argument('--disable-dev-shm-usage')
        try:
            driver = webdriver.Chrome(options=chrome_options)
        except WebDriverException as e:
            step = 'start ChromeDriver'
            raise RuntimeError(f'Could not start ChromeDriver: {str(e)}')
        try:
            step = 'selenium get url'
            driver.get(url)
            time.sleep(5)
            # Loop through all iframes and use the first with a valid, non-empty src
            body_text = ''
            iframe_srcs = []
            valid_iframe = False
            iframes = driver.find_elements(By.TAG_NAME, "iframe")
            for iframe in iframes:
                src = iframe.get_attribute("src")
                iframe_srcs.append(src)
                if src and isinstance(src, str) and src.strip():
                    iframe_url = urljoin(url, src)
                    print(f"Trying iframe src: {iframe_url}")
                    try:
                        driver.get(iframe_url)
                        time.sleep(5)
                        body_text = driver.find_element(By.TAG_NAME, "body").text
                        valid_iframe = True
                        print(f"Used iframe src: {iframe_url}")
                        break
                    except Exception as e:
                        print(f"Failed to load iframe URL: {e}, trying next iframe.")
            if not valid_iframe:
                print(f"No valid iframe found. All iframe srcs: {iframe_srcs}. Falling back to main page scraping.")
                body_text = driver.find_element(By.TAG_NAME, "body").text
            # Truncate body_text to 10,000 characters for Gemini
            body_text = body_text[:10000]
        except Exception as e:
            step = 'selenium exception'
            driver.quit()
            raise RuntimeError(f'Failed to extract text with Selenium: {str(e)}')
        # Send to Gemini API
        step = 'prepare prompt for Gemini'
        prompt = (
            "Extract all job listings from the following text. "
            "For each job, return a JSON list of objects with keys: 'Job', 'Description', and 'Location'. "
            "If a field is missing, use an empty string. Only return the JSON list, nothing else.\n\n"
            f"Text:\n{body_text}"
        )
        payload = {
            "contents": [{"parts": [{"text": prompt}]}]
        }
        headers = {"Content-Type": "application/json"}
        jobs = []
        try:
            step = 'call Gemini'
            response = http_requests.post(GEMINI_URL, headers=headers, data=json.dumps(payload))
            if response.status_code != 200:
                print(f"Gemini API error: {response.text}")
                raise Exception(f"Gemini API error: {response.text}")
            gemini_data = response.json()
            try:
                model_text = gemini_data['candidates'][0]['content']['parts'][0]['text']
            except Exception:
                raise Exception("Could not parse Gemini API response.")
            # Try to extract JSON from the model_text
            try:
                match = re.search(r'\[.*\]', model_text, re.DOTALL)
                if match:
                    jobs = json.loads(match.group(0))
                else:
                    jobs = json.loads(model_text)
            except Exception:
                jobs = []
        except Exception as e:
            step = 'parse Gemini response'
            print('Gemini error:', traceback.format_exc())
            jobs = []
        # Fallback: extract jobs from HTML if Gemini fails or returns 500
        if not jobs:
            try:
                jobs = extract_jobs_from_html(driver)
            except Exception as e:
                print('Fallback HTML extraction error:', traceback.format_exc())
                jobs = []
        driver.quit()
        # Write to Excel
        try:
            step = 'write to Excel'
            wb = None
            ws = None
            if os.path.exists(EXCEL_PATH):
                try:
                    wb = openpyxl.load_workbook(EXCEL_PATH)
                    ws = wb.active
                except (BadZipFile, Exception):
                    # Corrupted or not a real Excel file, delete and recreate
                    os.remove(EXCEL_PATH)
                    wb = openpyxl.Workbook()
                    ws = wb.active
                    if ws is not None:
                        ws.append(['Job', 'Description', 'Location'])
            else:
                wb = openpyxl.Workbook()
                ws = wb.active
                if ws is not None:
                    ws.append(['Job', 'Description', 'Location'])
            # Remove old data (except header)
            if ws is not None and ws.max_row > 1:
                ws.delete_rows(2, ws.max_row - 1)
            if ws is not None:
                for job in jobs:
                    ws.append([
                        job.get('Job', ''),
                        job.get('Description', ''),
                        job.get('Location', '')
                    ])
                wb.save(EXCEL_PATH)
            else:
                raise RuntimeError('Excel worksheet is None')
        except Exception as e:
            step = 'Excel write error'
            print('Excel error:', traceback.format_exc())
            raise RuntimeError(f'Failed to write to Excel: {str(e)}')
        return jsonify({'success': True, 'jobs': jobs})
    except Exception as e:
        print(f'ERROR at step: {step}\n', traceback.format_exc())
        return jsonify({'error': f'Error at step: {step}: {str(e)}'}), 500

if __name__ == '__main__':
    app.run(debug=True) 